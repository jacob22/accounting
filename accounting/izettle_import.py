#!/usr/bin/env python

# Copyright 2019 Open End AB
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#    http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import argparse
import mmap
import xlrd
import sys
import openpyxl
import re
import numbers
import decimal
import collections
import bson 
import locale

import pytransact
import pytransact.context
from pytransact import blm
import members

from accounting import db
from pytransact.commit import CommitContext, CallBlm, wait_for_commit
import blm.accounting

import sys
if sys.version_info < (3,0,0):
    PYT3 = False
else:
    PYT3 = True
    import codecs

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('-o', '--org')
    parser.add_argument('-p', '--products')
    parser.add_argument('-s', '--statement')
    parser.add_argument('-t', '--transactions')
    args = parser.parse_args()

    database = db.connect()
    with pytransact.context.ReadonlyContext(database):
        try:
            org, = blm.accounting.Org._query(id=args.org).run()
        except ValueError:
            name = args.org.decode('utf8')
            # Many orgs may have the same name - oh, well, just pick one
            org = blm.accounting.Org._query(name=name).run()[0]

    if args.products:
        print('Importing iZettle products from %s' % args.products)
        import_products_file(org.id[0], args.products)

    if args.statement:
        print('Parsing statement: %s' % args.statement)
        parse_statement_xlsx(args.statement)

    if args.transactions:
        print('Importing iZettle transactions from %s' % args.transactions)
        import_transactions(org.id[0], args.transactions)


def parse_statement_xlsx(filename):
    wb = openpyxl.load_workbook(filename=filename, read_only=True)
    sheet, = wb.get_sheet_names()
    ws = wb[sheet]
    #print ws.max_row, ws.max_column

    # Check that sheet looks ok.
    header_is, = ws['A1:F1']
    # TODO: find english heading texts
    header_should_en = (u'Bokf\xf6rd', u'Betalningsdag', u'Kvitto', u'Typ', u'Belopp', u'Saldo')
    header_should_sv = (u'Bokf\xf6rd', u'Betalningsdag', u'Kvitto', u'Typ', u'Belopp', u'Saldo')
    for cell, header_text_en, header_text_sv in zip(header_is, header_should_en, header_should_sv):
        assert cell.font.b == True
        assert cell.value in (header_text_en, header_text_sv) 

    # Extract data.
    data_rows = []
    for row in ws.iter_rows(row_offset=1):
        row_values = []
        for cell in row:
            row_values.append(cell.value)
        data_rows.append(row_values)
        #print row_values
        #print (u' ; '.join([unicode(c) for c in row_values])).encode('utf-8')

    # Separate data.
    statement = []
    payment_to_bank_account = []
    for row in data_rows:
        (date_entered, date_transfer, receipt, type, amount, balance) = row
        # Amount and balance are strings like u'23 250,18'.
        amount = decimal.Decimal(amount.replace(' ', '').replace(',', '.'))
        balance = decimal.Decimal(balance.replace(' ', '').replace(',', '.'))
        try:
            receipt = int(receipt)
        except TypeError:
            if date_transfer is None and receipt is None:
                payment_to_bank_account.append([date_entered, type, amount, balance])
                continue
            else:
                print('ERROR:', repr(row))

        statement.append([date_entered, date_transfer, receipt, type, amount, balance])
        #print (u' ; '.join([unicode(c) for c in row_values])).encode('utf-8')

    # Assert that data looks ok.
    re_date = re.compile(r'20\d\d-[0-1]\d-[0-3]\d')
    for row in statement:
        (date_entered, date_transfer, receipt, type, amount, balance) = row
        assert re_date.match(date_entered)
        assert re_date.match(date_transfer)
        assert len(type) > 1
        #print (u' | '.join([unicode(val) for val in row])).encode('utf-8')
        #print repr(row)
    return statement


def parse_transactions_xls(filename):
    with open(filename, 'rb') as f:
        xlsstring = f.read()
    transactions, paybacks = parse_transactions(xlsstring)
    return transactions, paybacks

def parse_transactions(xlsstring):

    if PYT3:
        xlsstring= codecs.decode(xlsstring.encode(), 'base64')

    wb = xlrd.open_workbook(file_contents=xlsstring)
    sheet, = wb.sheets()
    #print 'Sheet:',sheet.name, sheet.ncols, sheet.nrows

    first_data_row = 18
    last_data_row = sheet.nrows - 3

    # Search sheet to find start and end of data fields
    for nrow in range(sheet.nrows):
        cell = sheet.cell(nrow, 0).value
        if cell == u'Datum':
            first_data_row = nrow + 1
            break

    # Print entire sheet
    #for row in range(sheet.nrows):
    #    row_values = []
    #    for col in range(sheet.ncols):
    #        row_values.append(sheet.cell(row, col).value)
    #    print (u' ; '.join([unicode(val) for val in row_values])).encode('utf-8')

    header_row = first_data_row - 1

    # TODO: find english heading texts
    header_should_en = (u'Datum', u'Tid', u'Kvittonummer', u'Totalt', u'Avgift', u'Netto', 
        u'Betals\xe4tt', u'Korttyp', u'Sista siffror', u'Personal', u'Enhet',  u'Beskrivning')
    header_should_sv = (u'Datum', u'Tid', u'Kvittonummer', u'Totalt', u'Avgift', u'Netto', 
        u'Betals\xe4tt', u'Korttyp', u'Sista siffror', u'Personal', u'Enhet',  u'Beskrivning')
    for col in range(len(header_should_en)):
        cell = sheet.cell(header_row, col).value
        assert cell in (header_should_en[col], header_should_sv[col])
    # First row of footer should be empty
    assert sheet.cell(last_data_row + 1, 0).value == u''

    # Extract table data
    data_rows = []
    for row in range(first_data_row, last_data_row):
        row_values = []
        for col in range(sheet.ncols):
            row_values.append(sheet.cell(row, col).value)
        data_rows.append(row_values)
        #print row_values
        #print (u' | '.join([unicode(val) for val in row_values])).encode('utf-8')

    # Extract transactions, as lines with other format can appear in the middle of the sheet
    transactions = []
    paybacks = []

    Transaction = collections.namedtuple(
        'Transaction', '''transaction_date transaction_time receipt_number amount 
         izettle_fee netto payment_class_text card_type last_digits
         cashier device description'''.split()
    )
    Payback = collections.namedtuple(
        'Payback', '''transaction_date transaction_time amount 
         transaction_type timespan'''.split()
    )

    
    for row in data_rows:
        transaction = Transaction(*row[:12])._asdict()
        # Convert data
        if transaction['card_type'] == '---':
            transaction['payment_class'] = 'cash'
        else:
            transaction['payment_class'] = 'card'

        # TODO: English version of Retur
        if transaction['payment_class_text'] in ('Refund', 'Retur'):
            transaction['is_return'] = True
        else:
            transaction['is_return'] = False
        try:
            receipt_number = int(transaction['receipt_number'])
        except ValueError: 
            if transaction['receipt_number'] == u'' and len(row[-1]) > 0:
                # Lines with another table format. 
                payback = Payback(row[0], row[1], row[6], row[7], row[12])._asdict()
                paybacks.append(payback)
                continue
            else:
                print('ERROR: ', repr(row))
        transactions.append(transaction)
     
    # Check that transactions look ok.
    #re_date = re.compile(r'20\d\d-[0-1]\d-[0-3]\d')
    #re_time = re.compile(r'[012]\d:\d\d:\d\d')
    #re_last_digits = re.compile(r'[\d]{4}')

    #for t in transactions:
    #    try:
    #        if t.card_type == '---':
    #            # Cash payment.
    #            pass
    #        assert re_date.match(t.transaction_date)
    #        assert re_time.match(t.transaction_time)
    #        assert isinstance(t.amount, numbers.Number)
    #        assert isinstance(t.izettle_fee, numbers.Number)
    #        assert isinstance(t.netto, numbers.Number)
    #    except AssertionError:
    #        print 'ERROR:', repr(t)
    #    print ' | '.join([unicode(c) for c in t.values()])
    return transactions, paybacks


def parse_products_xls(file):
    try:
        file.read
    except AttributeError:
        file = open(file)

    xlsstring = file.read()
    file.close()
    if PYT3:
        xlsstring= codecs.decode(xlsstring.encode(), 'base64')

    wb = xlrd.open_workbook(file_contents=xlsstring)
    sheet, = wb.sheets()

    # Print entire sheet
    #for row in range(sheet.nrows):
    #    row_values = []
    #    for col in range(sheet.ncols):
    #        row_values.append(sheet.cell(row, col).value)
    #    print (u' ; '.join([unicode(val) for val in row_values])).encode('utf-8')

    header_row = 0
    header_should_en = [u'Name', u'Variant name', u'Price', u'Barcode', 
        u'Custom unit', u'VAT in %', u'Product Id']
    header_should_sv_old = [u'Namn', u'Namn p\xe5 variant', u'Pris', u'Streckkod', 
        u'Annan enhet (max 4 tecken)', u'Moms (i %)', u'Produkt-ID']
    header_should_sv = [u'Namn', u'Namn p\xe5 variant', u'Pris', u'Streckkod', 
        u'Annan enhet (max 4 tecken)', u'Moms (i %)', u'Product Id']

    for col in range(len(header_should_en)):
        cell = sheet.cell(header_row, col).value
        if cell == header_should_en[col] or cell == header_should_sv[col] or cell == header_should_sv_old[col]:
            pass
        else:
            raise ValueError('Bad column: %r' % cell)

    first_data_row = 1 
    last_data_row = sheet.nrows

    # Extract table data
    data_rows = []
    for row in range(first_data_row, last_data_row):
        row_values = []
        for col in range(sheet.ncols):
            row_values.append(sheet.cell(row, col).value)
        data_rows.append(row_values)
        #print row_values
        #print (u' | '.join([unicode(val) for val in row_values])).encode('utf-8')
    
    keys = ('name', 'variant', 'izPrice', 'barcode', 'customUnit', 'vatPercentage', 'productId')
    for row in data_rows:
        (name, variant, izPrice, barcode, customUnit, vatPercentage, productId) = row
        if izPrice:
            izPrice = decimal.Decimal(izPrice.replace(' ', '').replace(',', '.'))
        else:
            izPrice = []
        vatPercentage = decimal.Decimal(vatPercentage)
        row = (name, variant, izPrice, barcode, customUnit, vatPercentage, productId)
        product = dict(zip(keys, row))
        yield product
        

def import_products(org, products):
    import blm.members
    database = db.connect()
    interested = 'izettle-import-%s' % bson.objectid.ObjectId()
    with pytransact.commit.CommitContext(database) as ctx:
        ctx.setMayChange(True)
        op = CallBlm('members', 'import_izettle_products', [[org], [products]])
        ctx.runCommit([op], interested=interested)
        result, errors = pytransact.commit.wait_for_commit(database, interested=interested)
        assert not errors, errors

def import_products_file(org, filename):
    import blm.members
    with open(filename) as f:
        products = f.read().encode('base64')

    database = db.connect()
    interested = 'izettle-import-%s' % bson.objectid.ObjectId()
    with pytransact.commit.CommitContext(database) as ctx:
        ctx.setMayChange(True)
        op = CallBlm('members', 'import_izettle_products_file', [[org], [products]])
        ctx.runCommit([op], interested=interested)
        result, errors = pytransact.commit.wait_for_commit(database, interested=interested)
        assert not errors, errors

def import_transactions(org, transactionsfilename):
    with open(transactionsfilename) as f:
        filedata = f.read().encode('base64')
    import blm.members
    database = db.connect()
    interested = 'izettle-import-%s' % bson.objectid.ObjectId()
    with pytransact.commit.CommitContext(database) as ctx:
        ctx.setMayChange(True)
        op = CallBlm('members', 'import_izettle_payments', [[org], [filedata]])
        ctx.runCommit([op], interested=interested)
        result, errors = pytransact.commit.wait_for_commit(database, interested=interested)
        assert not errors, errors


def correlate(statement, transactions):
    need_manual_attention = []
    parsable = []
    d_statement = collections.defaultdict(list)
    for s in statement:
        receipt = s[2]
        d_statement[receipt].append(s)
    d_statement = dict(d_statement) # Turn of defaultdict behaviour
    for transaction in transactions:
        (date, time, receipt, amount, fee, netto, payment_class, card_type, last_digits,
            cashier, device, description) = transaction
        shoppinglist = interpret_description(description)
        if not shoppinglist:
            #print 'ERROR:', description
            need_manual_attention.append(transaction)
            continue
        price_sum = 0
        for nr, item in shoppinglist:
            price_sum += nr * item['price']
        assert price_sum == round(amount, 4)
        if receipt in d_statement:
            print('TRANSACTION:', transaction)
            for s in d_statement[receipt]:
                print('STATEMENT:', s)
            print('--------')

if __name__ == '__main__': 
    main()
